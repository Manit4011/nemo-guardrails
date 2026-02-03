import os
import boto3
import json
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from app import create_embeddings, call_llm
from analyze import group_similar_questions
   
def get_precision_prompt(row):
    prompt = f"""
Evaluate the following response based on these metrics: precision, answer relevance, and context recall.

**Question:** {row['rewritten_query']}
**Response:** {row['response']}
**Contexts:** {row['reference'] if row['reference'] else "No context available."}

Please provide scores for each metric in JSON format as shown below:
{{
    "Accuracy": "<How well does the response answer the user question? Score between 0 and 1>",
    "Precision": "<How well does the answer use the given context? Score between 0 and 1>",
    "Answer_relevance": "<Is the answer relevant to the question? Score between 0 and 1>",
    "Context_recall": "<How well does the response recall information from the context? Score between 0 and 1>",
    "Reason": "<Brief explanation why you think the response deserves this score (max 20 words)>"
}}

Return only the output JSON without any additional text.
"""
    return prompt

def get_accuracy(df):
    df['response'].fillna("", inplace=True)  # Avoid chaining warnings
    scores = {
        'accuracy': [],
        'precision': [], 
        'answer_relevance': [], 
        'context_recall': [], 
        'reason': []
    }
    for index, row in df.iterrows():
        client = boto3.client(service_name='bedrock-runtime', region_name="ap-south-1")
        eval_prompt = get_precision_prompt(row)
        scores_response = call_llm(client, eval_prompt, model='meta.llama3-70b-instruct-v1:0')
        print(scores_response)
        try:
            evaluation_scores = json.loads(scores_response)
            print(evaluation_scores)
            for key in scores.keys():
                scores[key].append(evaluation_scores.get(key.capitalize(), None))
        except json.JSONDecodeError:
            print(f"Error: Failed to parse evaluation scores for index {index}")
    for key, value in scores.items():
        df[key] = value
    df['accuracy'] = pd.to_numeric(df['accuracy'], errors='coerce')
    return df

def style_excel(writer, df, inaccurate_df):
    workbook = writer.book
    
    # Styling for 'Accuracy analysis' sheet
    data = writer.sheets['Accuracy analysis']
    low_acc = writer.sheets['Inaccurate Responses']
    
    # Set column width for both sheets
    for sheet in [data, low_acc]:
        for idx, col in enumerate(df.columns, 1):
            sheet.column_dimensions[chr(64 + idx)].width = 25  # Set column width
    
    # Header style: bold, center aligned, larger font, colored background
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    
    for sheet in [data, low_acc]:
        for cell in sheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply conditional formatting for accuracy values less than 0.7 in red
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    data.conditional_formatting.add(
        f'E2:E{len(df)+1}',
        CellIsRule(operator='lessThan', formula=['0.7'], stopIfTrue=True, fill=red_fill)
    )
    
    # Apply alternating row colors for readability (zebra stripes effect)
    row_fill_1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for sheet in [data, low_acc]:
        for row in sheet.iter_rows(min_row=2, max_row=len(df) + 1):
            for cell in row:
                cell.fill = row_fill_1 if cell.row % 2 == 0 else row_fill_2
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column width based on content
    for sheet in [data, low_acc]:
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Adding extra buffer space
            sheet.column_dimensions[col_letter].width = adjusted_width


def analyze_accuracy(df=None, file_name=None, save_file="analysis.xlsx"):
    if file_name:
        df = pd.read_excel(file_name)
    
    df = df[df['response'].notna() & (df['response'] != "")]
    df = group_similar_questions(df)
    
    # Count top unique grouped queries
    top_grouped_queries = df['grouped_query'].value_counts().head(40).index.tolist()
    custom_df = df[df['grouped_query'].isin(top_grouped_queries)].copy()
    custom_df = custom_df.drop_duplicates(subset='grouped_query', keep='first')
    
    custom_df = get_accuracy(custom_df)
    inaccurate_responses = custom_df[custom_df['accuracy'] < 0.7][['rewritten_query','response','metadata','reference','accuracy','reason']].sort_values(by='accuracy', ascending=True)
    
    # Save analysis to Excel and apply styling
    with pd.ExcelWriter(save_file, engine='openpyxl', mode='w') as writer:
        custom_df.to_excel(writer, sheet_name='Accuracy analysis', index=False)
        inaccurate_responses.to_excel(writer, sheet_name='Inaccurate Responses', index=False)
        style_excel(writer, custom_df, inaccurate_responses)
    
    print(f"Analysis successfully saved to {save_file}")
    return custom_df

if __name__ == '__main__':
    df = analyze_accuracy(file_name='conv_history.xlsx')